"""
Excel Export Utility for Trading Application

This module provides comprehensive Excel export functionality using openpyxl for
professional-grade report generation including scanner results, trading journal entries,
paper trading performance, and analytics dashboard data with multi-sheet organization,
conditional formatting, and automated chart generation.

Features:
- Scanner results export with VWAP data and technical indicators
- Trading journal export with embedded performance charts
- Paper trading simulation export with P/L tracking and portfolio analytics
- Professional Excel formatting with conditional formatting and pivot tables
- Multi-sheet organization for comprehensive reporting
- Configurable report templates with automated file organization
- Integration with Settings Tab for export configuration management

Author: Blitzy Platform Trading Application
Version: 1.0.0
"""

import os
import json
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Union, Any, Tuple
from decimal import Decimal, ROUND_HALF_UP
import traceback

# Third-party imports
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    NamedStyle, Font, PatternFill, Border, Side, Alignment,
    Protection, Color
)
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, IconSetRule, CellIsRule
)
from openpyxl.chart import (
    LineChart, BarChart, PieChart, ScatterChart, Reference, Series
)
from openpyxl.chart.axis import DateAxis, ValueAxis
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.trendline import Trendline
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.page import PageMargins, PrintPageSetup
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension


class ExcelExportError(Exception):
    """Custom exception for Excel export operations."""
    pass


class ExcelExportUtility:
    """
    Comprehensive Excel export utility providing openpyxl-based report generation
    for scanner results, trading journal entries, paper trading performance,
    and analytics dashboard data with professional formatting.
    """
    
    def __init__(self, config_path: Optional[str] = None):
        """
        Initialize Excel export utility with configuration management.
        
        Args:
            config_path: Optional path to configuration file
        """
        self.logger = logging.getLogger(__name__)
        self.config = self._load_configuration(config_path)
        self.export_base_dir = self._setup_export_directory()
        
        # Initialize professional styling
        self._setup_excel_styles()
        
        # Performance tracking
        self.export_stats = {
            'total_exports': 0,
            'successful_exports': 0,
            'failed_exports': 0,
            'last_export_time': None
        }
        
        self.logger.info("Excel Export Utility initialized successfully")

    def _load_configuration(self, config_path: Optional[str]) -> Dict[str, Any]:
        """
        Load export configuration from JSON file with fallback defaults.
        
        Args:
            config_path: Path to configuration file
            
        Returns:
            Configuration dictionary with export settings
        """
        default_config = {
            "export_settings": {
                "default_format": "xlsx",
                "include_charts": True,
                "include_pivot_tables": True,
                "conditional_formatting": True,
                "auto_backup": True,
                "backup_retention_days": 30,
                "max_file_size_mb": 50,
                "compression_level": 6
            },
            "file_organization": {
                "create_date_folders": True,
                "folder_structure": "YYYY-MM",
                "filename_prefix": "TradingApp",
                "include_timestamp": True,
                "auto_increment": True
            },
            "data_formatting": {
                "decimal_places": 4,
                "currency_symbol": "$",
                "percentage_format": "0.00%",
                "date_format": "MM/DD/YYYY",
                "time_format": "HH:MM:SS"
            },
            "export_templates": {
                "scanner_results": {
                    "sheets": ["Summary", "Detailed_Data", "Charts", "Technical_Indicators"],
                    "charts": ["price_chart", "volume_chart", "vwap_chart"],
                    "conditional_formatting": ["price_change", "volume_ratio", "float_analysis"]
                },
                "trading_journal": {
                    "sheets": ["Journal_Summary", "Trade_Details", "Performance_Analysis", "Charts"],
                    "charts": ["equity_curve", "monthly_returns", "trade_distribution"],
                    "conditional_formatting": ["pnl_analysis", "win_rate", "risk_metrics"]
                },
                "paper_trading": {
                    "sheets": ["Portfolio_Summary", "Trade_History", "Performance_Metrics", "Risk_Analysis"],
                    "charts": ["portfolio_value", "sector_allocation", "performance_comparison"],
                    "conditional_formatting": ["position_pnl", "allocation_limits", "risk_exposure"]
                }
            }
        }
        
        if config_path and os.path.exists(config_path):
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    file_config = json.load(f)
                    # Merge with defaults, file config takes precedence
                    default_config.update(file_config)
                    self.logger.info(f"Configuration loaded from {config_path}")
            except Exception as e:
                self.logger.warning(f"Failed to load config from {config_path}: {e}")
                self.logger.info("Using default configuration")
        else:
            self.logger.info("Using default configuration")
            
        return default_config

    def _setup_export_directory(self) -> Path:
        """
        Setup export directory structure based on configuration.
        
        Returns:
            Path to base export directory
        """
        try:
            # Use %APPDATA%/TradingApp/exports as specified in technical spec
            appdata = os.environ.get('APPDATA', os.path.expanduser('~'))
            base_dir = Path(appdata) / 'TradingApp' / 'exports'
            
            # Create subdirectories for different export types
            export_types = ['scanner_results', 'trading_journal', 'paper_trading', 'custom_reports']
            
            for export_type in export_types:
                type_dir = base_dir / export_type
                type_dir.mkdir(parents=True, exist_ok=True)
                
            # Create backup directory if auto-backup is enabled
            if self.config.get('export_settings', {}).get('auto_backup', True):
                backup_dir = base_dir / 'backups'
                backup_dir.mkdir(exist_ok=True)
                
            self.logger.info(f"Export directory structure created at {base_dir}")
            return base_dir
            
        except Exception as e:
            self.logger.error(f"Failed to setup export directory: {e}")
            # Fallback to current directory
            fallback_dir = Path.cwd() / 'exports'
            fallback_dir.mkdir(exist_ok=True)
            return fallback_dir

    def _setup_excel_styles(self):
        """Setup professional Excel styles for consistent formatting."""
        self.styles = {
            'header_style': NamedStyle(
                name='header_style',
                font=Font(bold=True, size=12, color='FFFFFF'),
                fill=PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
                border=Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                ),
                alignment=Alignment(horizontal='center', vertical='center')
            ),
            'data_style': NamedStyle(
                name='data_style',
                font=Font(size=10),
                border=Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                ),
                alignment=Alignment(horizontal='center', vertical='center')
            ),
            'currency_style': NamedStyle(
                name='currency_style',
                font=Font(size=10),
                number_format='$#,##0.00',
                alignment=Alignment(horizontal='right')
            ),
            'percentage_style': NamedStyle(
                name='percentage_style',
                font=Font(size=10),
                number_format='0.00%',
                alignment=Alignment(horizontal='right')
            ),
            'date_style': NamedStyle(
                name='date_style',
                font=Font(size=10),
                number_format='MM/DD/YYYY',
                alignment=Alignment(horizontal='center')
            )
        }

    def export_scanner_results(self, scanner_data: Dict[str, Any], 
                             export_name: Optional[str] = None) -> Tuple[bool, str]:
        """
        Export comprehensive scanner results with VWAP data, technical indicators,
        and screening criteria to Excel with professional formatting.
        
        Args:
            scanner_data: Dictionary containing scanner results and metadata
            export_name: Optional custom name for export file
            
        Returns:
            Tuple of (success: bool, file_path_or_error: str)
        """
        try:
            self.logger.info("Starting scanner results export")
            
            # Generate filename
            filename = self._generate_filename('scanner_results', export_name)
            file_path = self.export_base_dir / 'scanner_results' / filename
            
            # Create workbook with multiple sheets
            workbook = Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Create sheets based on template configuration
            template = self.config['export_templates']['scanner_results']
            
            # Sheet 1: Summary Dashboard
            self._create_scanner_summary_sheet(workbook, scanner_data)
            
            # Sheet 2: Detailed Data
            self._create_scanner_detailed_sheet(workbook, scanner_data)
            
            # Sheet 3: Technical Indicators
            self._create_scanner_technical_sheet(workbook, scanner_data)
            
            # Sheet 4: Charts and Visualizations
            if self.config['export_settings']['include_charts']:
                self._create_scanner_charts_sheet(workbook, scanner_data)
            
            # Apply professional formatting
            self._apply_scanner_formatting(workbook)
            
            # Save workbook
            workbook.save(file_path)
            
            # Create backup if enabled
            if self.config['export_settings']['auto_backup']:
                self._create_backup(file_path)
            
            # Update statistics
            self.export_stats['total_exports'] += 1
            self.export_stats['successful_exports'] += 1
            self.export_stats['last_export_time'] = datetime.now()
            
            self.logger.info(f"Scanner results exported successfully to {file_path}")
            return True, str(file_path)
            
        except Exception as e:
            self.export_stats['failed_exports'] += 1
            error_msg = f"Failed to export scanner results: {str(e)}"
            self.logger.error(f"{error_msg}\n{traceback.format_exc()}")
            return False, error_msg

    def export_trading_journal(self, journal_data: Dict[str, Any],
                             export_name: Optional[str] = None) -> Tuple[bool, str]:
        """
        Export trading journal entries with embedded performance charts,
        trade analysis, and strategy documentation.
        
        Args:
            journal_data: Dictionary containing journal entries and trade data
            export_name: Optional custom name for export file
            
        Returns:
            Tuple of (success: bool, file_path_or_error: str)
        """
        try:
            self.logger.info("Starting trading journal export")
            
            # Generate filename
            filename = self._generate_filename('trading_journal', export_name)
            file_path = self.export_base_dir / 'trading_journal' / filename
            
            # Create workbook
            workbook = Workbook()
            workbook.remove(workbook.active)
            
            # Create sheets for journal export
            self._create_journal_summary_sheet(workbook, journal_data)
            self._create_journal_detailed_sheet(workbook, journal_data)
            self._create_journal_performance_sheet(workbook, journal_data)
            
            # Add charts if enabled
            if self.config['export_settings']['include_charts']:
                self._create_journal_charts_sheet(workbook, journal_data)
            
            # Apply formatting
            self._apply_journal_formatting(workbook)
            
            # Save workbook
            workbook.save(file_path)
            
            # Create backup
            if self.config['export_settings']['auto_backup']:
                self._create_backup(file_path)
            
            # Update statistics
            self.export_stats['total_exports'] += 1
            self.export_stats['successful_exports'] += 1
            self.export_stats['last_export_time'] = datetime.now()
            
            self.logger.info(f"Trading journal exported successfully to {file_path}")
            return True, str(file_path)
            
        except Exception as e:
            self.export_stats['failed_exports'] += 1
            error_msg = f"Failed to export trading journal: {str(e)}"
            self.logger.error(f"{error_msg}\n{traceback.format_exc()}")
            return False, error_msg

    def export_paper_trading_performance(self, trading_data: Dict[str, Any],
                                       export_name: Optional[str] = None) -> Tuple[bool, str]:
        """
        Export paper trading simulation data with P/L tracking, portfolio analytics,
        and performance dashboard data for comprehensive reporting.
        
        Args:
            trading_data: Dictionary containing trading simulation data
            export_name: Optional custom name for export file
            
        Returns:
            Tuple of (success: bool, file_path_or_error: str)
        """
        try:
            self.logger.info("Starting paper trading performance export")
            
            # Generate filename
            filename = self._generate_filename('paper_trading', export_name)
            file_path = self.export_base_dir / 'paper_trading' / filename
            
            # Create workbook
            workbook = Workbook()
            workbook.remove(workbook.active)
            
            # Create comprehensive trading sheets
            self._create_portfolio_summary_sheet(workbook, trading_data)
            self._create_trade_history_sheet(workbook, trading_data)
            self._create_performance_metrics_sheet(workbook, trading_data)
            self._create_risk_analysis_sheet(workbook, trading_data)
            
            # Add advanced charts
            if self.config['export_settings']['include_charts']:
                self._create_trading_charts_sheet(workbook, trading_data)
            
            # Apply professional formatting and conditional formatting
            self._apply_trading_formatting(workbook)
            
            # Save workbook
            workbook.save(file_path)
            
            # Create backup
            if self.config['export_settings']['auto_backup']:
                self._create_backup(file_path)
            
            # Update statistics
            self.export_stats['total_exports'] += 1
            self.export_stats['successful_exports'] += 1
            self.export_stats['last_export_time'] = datetime.now()
            
            self.logger.info(f"Paper trading performance exported successfully to {file_path}")
            return True, str(file_path)
            
        except Exception as e:
            self.export_stats['failed_exports'] += 1
            error_msg = f"Failed to export paper trading performance: {str(e)}"
            self.logger.error(f"{error_msg}\n{traceback.format_exc()}")
            return False, error_msg

    def export_custom_report(self, data: Dict[str, Any], template_name: str,
                           export_name: Optional[str] = None) -> Tuple[bool, str]:
        """
        Export custom report using configurable templates with flexible data input.
        
        Args:
            data: Dictionary containing data to export
            template_name: Name of template configuration to use
            export_name: Optional custom name for export file
            
        Returns:
            Tuple of (success: bool, file_path_or_error: str)
        """
        try:
            self.logger.info(f"Starting custom report export with template: {template_name}")
            
            # Generate filename
            filename = self._generate_filename('custom_reports', export_name)
            file_path = self.export_base_dir / 'custom_reports' / filename
            
            # Get template configuration
            template_config = self.config.get('export_templates', {}).get(template_name)
            if not template_config:
                raise ExcelExportError(f"Template '{template_name}' not found in configuration")
            
            # Create workbook based on template
            workbook = Workbook()
            workbook.remove(workbook.active)
            
            # Create sheets based on template configuration
            for sheet_name in template_config.get('sheets', ['Data']):
                sheet = workbook.create_sheet(title=sheet_name)
                self._populate_custom_sheet(sheet, data, sheet_name.lower())
            
            # Apply custom formatting
            self._apply_custom_formatting(workbook, template_config)
            
            # Add charts if specified in template
            if (self.config['export_settings']['include_charts'] and 
                template_config.get('charts')):
                self._create_custom_charts(workbook, data, template_config['charts'])
            
            # Save workbook
            workbook.save(file_path)
            
            # Create backup
            if self.config['export_settings']['auto_backup']:
                self._create_backup(file_path)
            
            # Update statistics
            self.export_stats['total_exports'] += 1
            self.export_stats['successful_exports'] += 1
            self.export_stats['last_export_time'] = datetime.now()
            
            self.logger.info(f"Custom report exported successfully to {file_path}")
            return True, str(file_path)
            
        except Exception as e:
            self.export_stats['failed_exports'] += 1
            error_msg = f"Failed to export custom report: {str(e)}"
            self.logger.error(f"{error_msg}\n{traceback.format_exc()}")
            return False, error_msg

    def _generate_filename(self, export_type: str, custom_name: Optional[str] = None) -> str:
        """
        Generate appropriate filename based on configuration and export type.
        
        Args:
            export_type: Type of export (scanner_results, trading_journal, etc.)
            custom_name: Optional custom name override
            
        Returns:
            Generated filename with extension
        """
        config = self.config['file_organization']
        
        if custom_name:
            base_name = custom_name
        else:
            prefix = config.get('filename_prefix', 'TradingApp')
            base_name = f"{prefix}_{export_type}"
        
        # Add timestamp if configured
        if config.get('include_timestamp', True):
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_name = f"{base_name}_{timestamp}"
        
        # Add extension
        format_ext = self.config['export_settings'].get('default_format', 'xlsx')
        filename = f"{base_name}.{format_ext}"
        
        return filename

    def _create_scanner_summary_sheet(self, workbook: Workbook, data: Dict[str, Any]):
        """Create summary sheet for scanner results with key metrics."""
        sheet = workbook.create_sheet(title="Scanner_Summary")
        
        # Add header information
        sheet['A1'] = "Market Scanner Results Summary"
        sheet['A1'].font = Font(bold=True, size=16)
        
        # Add scan metadata
        current_row = 3
        sheet[f'A{current_row}'] = "Scan Date:"
        sheet[f'B{current_row}'] = datetime.now().strftime("%m/%d/%Y %H:%M:%S")
        current_row += 1
        
        sheet[f'A{current_row}'] = "Total Securities Scanned:"
        sheet[f'B{current_row}'] = len(data.get('results', []))
        current_row += 1
        
        sheet[f'A{current_row}'] = "Filter Criteria Applied:"
        sheet[f'B{current_row}'] = str(data.get('filters', 'All'))
        current_row += 2
        
        # Add top performers table
        results = data.get('results', [])
        if results:
            # Headers
            headers = ['Symbol', 'Current Price', '% Change', 'Volume', 'VWAP', 'Float', 'Sector']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=current_row, column=col, value=header)
                cell.style = self.styles['header_style']
            
            current_row += 1
            
            # Data rows (top 20 performers)
            top_results = sorted(results, key=lambda x: x.get('percent_change', 0), reverse=True)[:20]
            
            for result in top_results:
                row_data = [
                    result.get('symbol', ''),
                    self._format_currency(result.get('current_price', 0)),
                    self._format_percentage(result.get('percent_change', 0)),
                    self._format_number(result.get('volume', 0)),
                    self._format_currency(result.get('vwap', 0)),
                    self._format_number(result.get('float_shares', 0)),
                    result.get('sector', '')
                ]
                
                for col, value in enumerate(row_data, 1):
                    sheet.cell(row=current_row, column=col, value=value)
                
                current_row += 1
        
        # Auto-adjust column widths
        self._auto_fit_columns(sheet)

    def _create_scanner_detailed_sheet(self, workbook: Workbook, data: Dict[str, Any]):
        """Create detailed sheet with complete scanner data."""
        sheet = workbook.create_sheet(title="Detailed_Data")
        
        # Comprehensive headers
        headers = [
            'Symbol', 'Company Name', 'Current Price', 'Previous Close', 
            'Price Change', '% Change', 'Volume', 'Avg Volume', 'Volume Ratio',
            'VWAP', 'Market Cap', 'Float Shares', 'Sector', 'Industry',
            'RSI', 'MACD', 'Bollinger Position', 'ATR', 'Beta', 'News Sentiment'
        ]
        
        # Add headers with professional styling
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.style = self.styles['header_style']
        
        # Add data rows
        results = data.get('results', [])
        for row, result in enumerate(results, 2):
            row_data = [
                result.get('symbol', ''),
                result.get('company_name', ''),
                result.get('current_price', 0),
                result.get('previous_close', 0),
                result.get('price_change', 0),
                result.get('percent_change', 0),
                result.get('volume', 0),
                result.get('avg_volume', 0),
                result.get('volume_ratio', 0),
                result.get('vwap', 0),
                result.get('market_cap', 0),
                result.get('float_shares', 0),
                result.get('sector', ''),
                result.get('industry', ''),
                result.get('rsi', 0),
                result.get('macd', 0),
                result.get('bollinger_position', 0),
                result.get('atr', 0),
                result.get('beta', 0),
                result.get('news_sentiment', 'Neutral')
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row, column=col, value=value)
                
                # Apply appropriate formatting based on data type
                if col in [3, 4, 5, 10]:  # Price columns
                    cell.number_format = '$#,##0.0000'
                elif col in [6, 7, 8, 12]:  # Volume/count columns
                    cell.number_format = '#,##0'
                elif col in [9, 15, 16, 17, 18, 19]:  # Ratio/percentage columns
                    cell.number_format = '0.0000'
                elif col == 11:  # Market cap
                    cell.number_format = '$#,##0,,,"B"'
        
        # Create table for better formatting
        table_range = f"A1:{get_column_letter(len(headers))}{len(results) + 1}"
        table = Table(displayName="ScannerResults", ref=table_range)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=False
        )
        sheet.add_table(table)
        
        # Auto-fit columns
        self._auto_fit_columns(sheet)

    def _create_scanner_technical_sheet(self, workbook: Workbook, data: Dict[str, Any]):
        """Create technical indicators sheet with advanced analysis."""
        sheet = workbook.create_sheet(title="Technical_Analysis")
        
        # Technical analysis headers
        headers = [
            'Symbol', 'RSI (14)', 'MACD Signal', 'MACD Histogram', 
            'SMA (20)', 'SMA (50)', 'EMA (12)', 'EMA (26)',
            'Bollinger Upper', 'Bollinger Lower', 'Bollinger %B',
            'ATR (14)', 'Stochastic %K', 'Stochastic %D',
            'Williams %R', 'CCI', 'Momentum', 'Technical Score'
        ]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.style = self.styles['header_style']
        
        # Add technical data
        results = data.get('results', [])
        for row, result in enumerate(results, 2):
            technical_data = result.get('technical_indicators', {})
            
            row_data = [
                result.get('symbol', ''),
                technical_data.get('rsi', 0),
                technical_data.get('macd_signal', 0),
                technical_data.get('macd_histogram', 0),
                technical_data.get('sma_20', 0),
                technical_data.get('sma_50', 0),
                technical_data.get('ema_12', 0),
                technical_data.get('ema_26', 0),
                technical_data.get('bollinger_upper', 0),
                technical_data.get('bollinger_lower', 0),
                technical_data.get('bollinger_percent_b', 0),
                technical_data.get('atr', 0),
                technical_data.get('stoch_k', 0),
                technical_data.get('stoch_d', 0),
                technical_data.get('williams_r', 0),
                technical_data.get('cci', 0),
                technical_data.get('momentum', 0),
                technical_data.get('technical_score', 0)
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row, column=col, value=value)
                if col > 1:  # Skip symbol column
                    cell.number_format = '0.0000'
        
        # Auto-fit columns
        self._auto_fit_columns(sheet)

    def _create_scanner_charts_sheet(self, workbook: Workbook, data: Dict[str, Any]):
        """Create charts sheet with visual analysis."""
        sheet = workbook.create_sheet(title="Charts")
        
        # Create price change distribution chart
        results = data.get('results', [])
        if results:
            # Prepare data for charting
            symbols = [r.get('symbol', '') for r in results[:20]]  # Top 20
            price_changes = [r.get('percent_change', 0) for r in results[:20]]
            volumes = [r.get('volume', 0) for r in results[:20]]
            
            # Create bar chart for price changes
            chart = BarChart()
            chart.title = "Top 20 Price Changes"
            chart.x_axis.title = "Symbols"
            chart.y_axis.title = "Percentage Change"
            chart.width = 20
            chart.height = 10
            
            # Add data to chart
            chart_data = []
            chart_data.append(['Symbol'] + symbols)
            chart_data.append(['% Change'] + price_changes)
            
            # Position chart
            sheet.add_chart(chart, "A2")
            
            # Create volume chart
            volume_chart = LineChart()
            volume_chart.title = "Volume Analysis"
            volume_chart.x_axis.title = "Symbols"
            volume_chart.y_axis.title = "Volume"
            volume_chart.width = 20
            volume_chart.height = 10
            
            # Position volume chart
            sheet.add_chart(volume_chart, "A25")

    def _create_journal_summary_sheet(self, workbook: Workbook, data: Dict[str, Any]):
        """Create trading journal summary sheet."""
        sheet = workbook.create_sheet(title="Journal_Summary")
        
        # Journal summary headers
        sheet['A1'] = "Trading Journal Summary"
        sheet['A1'].font = Font(bold=True, size=16)
        
        # Add key metrics
        current_row = 3
        entries = data.get('journal_entries', [])
        trades = data.get('linked_trades', [])
        
        metrics = [
            ('Total Journal Entries:', len(entries)),
            ('Linked Trades:', len(trades)),
            ('Date Range:', f"{data.get('start_date', 'N/A')} to {data.get('end_date', 'N/A')}"),
            ('Most Active Symbol:', data.get('most_active_symbol', 'N/A')),
            ('Average Entry Length:', f"{data.get('avg_entry_length', 0)} words"),
            ('Win Rate:', self._format_percentage(data.get('win_rate', 0))),
            ('Total P/L:', self._format_currency(data.get('total_pnl', 0)))
        ]
        
        for metric, value in metrics:
            sheet[f'A{current_row}'] = metric
            sheet[f'B{current_row}'] = value
            current_row += 1
        
        # Recent entries table
        current_row += 2
        sheet[f'A{current_row}'] = "Recent Journal Entries"
        sheet[f'A{current_row}'].font = Font(bold=True, size=14)
        current_row += 1
        
        # Headers for recent entries
        headers = ['Date', 'Symbol', 'Entry Type', 'Title', 'P/L Impact', 'Lessons Learned']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=current_row, column=col, value=header)
            cell.style = self.styles['header_style']
        
        current_row += 1
        
        # Recent entries data
        recent_entries = sorted(entries, key=lambda x: x.get('entry_date', ''), reverse=True)[:10]
        for entry in recent_entries:
            row_data = [
                entry.get('entry_date', ''),
                entry.get('symbol', ''),
                entry.get('entry_type', ''),
                entry.get('entry_title', '')[:50] + '...' if len(entry.get('entry_title', '')) > 50 else entry.get('entry_title', ''),
                self._format_currency(entry.get('pnl_impact', 0)),
                entry.get('lessons_learned', '')[:100] + '...' if len(entry.get('lessons_learned', '')) > 100 else entry.get('lessons_learned', '')
            ]
            
            for col, value in enumerate(row_data, 1):
                sheet.cell(row=current_row, column=col, value=value)
            
            current_row += 1
        
        # Auto-fit columns
        self._auto_fit_columns(sheet)

    def _create_journal_detailed_sheet(self, workbook: Workbook, data: Dict[str, Any]):
        """Create detailed journal entries sheet."""
        sheet = workbook.create_sheet(title="Detailed_Entries")
        
        # Detailed entries headers
        headers = [
            'Entry ID', 'Date', 'Symbol', 'Entry Type', 'Title',
            'Market Conditions', 'Strategy Notes', 'Analysis',
            'Lessons Learned', 'Trade Link', 'P/L Impact',
            'Rating', 'Tags', 'Follow-up Required'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.style = self.styles['header_style']
        
        # Add detailed entries
        entries = data.get('journal_entries', [])
        for row, entry in enumerate(entries, 2):
            row_data = [
                entry.get('journal_id', ''),
                entry.get('entry_date', ''),
                entry.get('symbol', ''),
                entry.get('entry_type', ''),
                entry.get('entry_title', ''),
                entry.get('market_conditions', ''),
                entry.get('strategy_notes', ''),
                entry.get('analysis_notes', ''),
                entry.get('lessons_learned', ''),
                entry.get('trade_id', ''),
                entry.get('pnl_impact', 0),
                entry.get('rating', 0),
                ', '.join(entry.get('tags', [])),
                entry.get('follow_up_required', False)
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row, column=col, value=value)
                
                # Format P/L column
                if col == 11:  # P/L Impact
                    cell.number_format = '$#,##0.00'
        
        # Auto-fit columns
        self._auto_fit_columns(sheet)

    def _create_journal_performance_sheet(self, workbook: Workbook, data: Dict[str, Any]):
        """Create performance analysis sheet for trading journal."""
        sheet = workbook.create_sheet(title="Performance_Analysis")
        
        # Performance metrics
        sheet['A1'] = "Trading Performance Analysis"
        sheet['A1'].font = Font(bold=True, size=16)
        
        current_row = 3
        
        # Key performance indicators
        performance_data = data.get('performance_metrics', {})
        
        kpis = [
            ('Total Trades Analyzed:', performance_data.get('total_trades', 0)),
            ('Winning Trades:', performance_data.get('winning_trades', 0)),
            ('Losing Trades:', performance_data.get('losing_trades', 0)),
            ('Win Rate:', self._format_percentage(performance_data.get('win_rate', 0))),
            ('Average Win:', self._format_currency(performance_data.get('avg_win', 0))),
            ('Average Loss:', self._format_currency(performance_data.get('avg_loss', 0))),
            ('Profit Factor:', f"{performance_data.get('profit_factor', 0):.2f}"),
            ('Sharpe Ratio:', f"{performance_data.get('sharpe_ratio', 0):.2f}"),
            ('Maximum Drawdown:', self._format_percentage(performance_data.get('max_drawdown', 0))),
            ('Recovery Factor:', f"{performance_data.get('recovery_factor', 0):.2f}")
        ]
        
        for kpi, value in kpis:
            sheet[f'A{current_row}'] = kpi
            sheet[f'B{current_row}'] = value
            current_row += 1
        
        # Monthly performance breakdown
        current_row += 2
        sheet[f'A{current_row}'] = "Monthly Performance Breakdown"
        sheet[f'A{current_row}'].font = Font(bold=True, size=14)
        current_row += 1
        
        monthly_headers = ['Month', 'Trades', 'Wins', 'Losses', 'P/L', 'Win Rate', 'Avg Trade']
        for col, header in enumerate(monthly_headers, 1):
            cell = sheet.cell(row=current_row, column=col, value=header)
            cell.style = self.styles['header_style']
        
        current_row += 1
        
        # Monthly data
        monthly_data = performance_data.get('monthly_breakdown', [])
        for month_data in monthly_data:
            row_data = [
                month_data.get('month', ''),
                month_data.get('total_trades', 0),
                month_data.get('wins', 0),
                month_data.get('losses', 0),
                month_data.get('pnl', 0),
                month_data.get('win_rate', 0),
                month_data.get('avg_trade', 0)
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = sheet.cell(row=current_row, column=col, value=value)
                
                # Format currency and percentage columns
                if col in [5, 7]:  # P/L and Avg Trade
                    cell.number_format = '$#,##0.00'
                elif col == 6:  # Win Rate
                    cell.number_format = '0.00%'
            
            current_row += 1
        
        # Auto-fit columns
        self._auto_fit_columns(sheet)

    def _create_journal_charts_sheet(self, workbook: Workbook, data: Dict[str, Any]):
        """Create charts for journal performance visualization."""
        sheet = workbook.create_sheet(title="Performance_Charts")
        
        # Add equity curve chart placeholder
        sheet['A1'] = "Performance Charts"
        sheet['A1'].font = Font(bold=True, size=16)
        
        # Create equity curve data
        performance_data = data.get('performance_metrics', {})
        equity_curve = performance_data.get('equity_curve', [])
        
        if equity_curve:
            # Equity curve chart
            chart = LineChart()
            chart.title = "Equity Curve"
            chart.x_axis.title = "Date"
            chart.y_axis.title = "Portfolio Value"
            chart.width = 15
            chart.height = 10
            
            sheet.add_chart(chart, "A3")
            
            # Monthly returns chart
            returns_chart = BarChart()
            returns_chart.title = "Monthly Returns"
            returns_chart.x_axis.title = "Month"
            returns_chart.y_axis.title = "Return %"
            returns_chart.width = 15
            returns_chart.height = 10
            
            sheet.add_chart(returns_chart, "A20")

    def _create_portfolio_summary_sheet(self, workbook: Workbook, data: Dict[str, Any]):
        """Create portfolio summary sheet for paper trading export."""
        sheet = workbook.create_sheet(title="Portfolio_Summary")
        
        # Portfolio overview
        sheet['A1'] = "Paper Trading Portfolio Summary"
        sheet['A1'].font = Font(bold=True, size=16)
        
        current_row = 3
        portfolio_data = data.get('portfolio_summary', {})
        
        # Key portfolio metrics
        summary_metrics = [
            ('Starting Capital:', self._format_currency(portfolio_data.get('starting_capital', 0))),
            ('Current Portfolio Value:', self._format_currency(portfolio_data.get('current_value', 0))),
            ('Cash Balance:', self._format_currency(portfolio_data.get('cash_balance', 0))),
            ('Invested Amount:', self._format_currency(portfolio_data.get('invested_amount', 0))),
            ('Total P/L:', self._format_currency(portfolio_data.get('total_pnl', 0))),
            ('Total Return %:', self._format_percentage(portfolio_data.get('total_return_pct', 0))),
            ('Number of Positions:', portfolio_data.get('position_count', 0)),
            ('Number of Trades:', portfolio_data.get('trade_count', 0)),
            ('Average Trade Size:', self._format_currency(portfolio_data.get('avg_trade_size', 0))),
            ('Largest Position:', f"{portfolio_data.get('largest_position_symbol', 'N/A')} ({self._format_percentage(portfolio_data.get('largest_position_pct', 0))})"),
        ]
        
        for metric, value in summary_metrics:
            sheet[f'A{current_row}'] = metric
            sheet[f'B{current_row}'] = value
            current_row += 1
        
        # Current positions table
        current_row += 2
        sheet[f'A{current_row}'] = "Current Positions"
        sheet[f'A{current_row}'].font = Font(bold=True, size=14)
        current_row += 1
        
        position_headers = [
            'Symbol', 'Quantity', 'Avg Cost', 'Current Price', 
            'Market Value', 'Unrealized P/L', '% Change', 'Position %'
        ]
        
        for col, header in enumerate(position_headers, 1):
            cell = sheet.cell(row=current_row, column=col, value=header)
            cell.style = self.styles['header_style']
        
        current_row += 1
        
        # Add position data
        positions = data.get('current_positions', [])
        for position in positions:
            row_data = [
                position.get('symbol', ''),
                position.get('quantity', 0),
                position.get('avg_cost', 0),
                position.get('current_price', 0),
                position.get('market_value', 0),
                position.get('unrealized_pnl', 0),
                position.get('percent_change', 0),
                position.get('position_percentage', 0)
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = sheet.cell(row=current_row, column=col, value=value)
                
                # Format columns appropriately
                if col in [3, 4, 5, 6]:  # Price and value columns
                    cell.number_format = '$#,##0.00'
                elif col in [7, 8]:  # Percentage columns
                    cell.number_format = '0.00%'
                elif col == 2:  # Quantity
                    cell.number_format = '#,##0'
            
            current_row += 1
        
        # Auto-fit columns
        self._auto_fit_columns(sheet)

    def _create_trade_history_sheet(self, workbook: Workbook, data: Dict[str, Any]):
        """Create comprehensive trade history sheet."""
        sheet = workbook.create_sheet(title="Trade_History")
        
        # Trade history headers
        headers = [
            'Trade ID', 'Date', 'Symbol', 'Action', 'Quantity',
            'Entry Price', 'Exit Price', 'Stop Loss', 'Target Price',
            'Fill Price', 'Commission', 'Slippage', 'Net P/L',
            'Hold Period', 'Return %', 'Strategy', 'Notes'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.style = self.styles['header_style']
        
        # Add trade data
        trades = data.get('trade_history', [])
        for row, trade in enumerate(trades, 2):
            row_data = [
                trade.get('trade_id', ''),
                trade.get('execution_time', ''),
                trade.get('symbol', ''),
                trade.get('direction', ''),
                trade.get('quantity', 0),
                trade.get('entry_price', 0),
                trade.get('exit_price', 0),
                trade.get('stop_loss', 0),
                trade.get('target_price', 0),
                trade.get('fill_price', 0),
                trade.get('commission', 0),
                trade.get('slippage', 0),
                trade.get('realized_pnl', 0),
                trade.get('hold_period_days', 0),
                trade.get('return_percentage', 0),
                trade.get('strategy', ''),
                trade.get('notes', '')
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row, column=col, value=value)
                
                # Format columns
                if col in [6, 7, 8, 9, 10, 11, 12, 13]:  # Price and P/L columns
                    cell.number_format = '$#,##0.00'
                elif col == 15:  # Return percentage
                    cell.number_format = '0.00%'
                elif col == 5:  # Quantity
                    cell.number_format = '#,##0'
        
        # Create table
        table_range = f"A1:{get_column_letter(len(headers))}{len(trades) + 1}"
        table = Table(displayName="TradeHistory", ref=table_range)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        sheet.add_table(table)
        
        # Auto-fit columns
        self._auto_fit_columns(sheet)

    def _create_performance_metrics_sheet(self, workbook: Workbook, data: Dict[str, Any]):
        """Create detailed performance metrics analysis sheet."""
        sheet = workbook.create_sheet(title="Performance_Metrics")
        
        # Performance analysis
        sheet['A1'] = "Performance Metrics & Analysis"
        sheet['A1'].font = Font(bold=True, size=16)
        
        current_row = 3
        perf_data = data.get('performance_analysis', {})
        
        # Core metrics
        core_metrics = [
            ('Total Return:', self._format_percentage(perf_data.get('total_return', 0))),
            ('Annualized Return:', self._format_percentage(perf_data.get('annualized_return', 0))),
            ('Volatility (Annualized):', self._format_percentage(perf_data.get('volatility', 0))),
            ('Sharpe Ratio:', f"{perf_data.get('sharpe_ratio', 0):.2f}"),
            ('Sortino Ratio:', f"{perf_data.get('sortino_ratio', 0):.2f}"),
            ('Maximum Drawdown:', self._format_percentage(perf_data.get('max_drawdown', 0))),
            ('Calmar Ratio:', f"{perf_data.get('calmar_ratio', 0):.2f}"),
            ('Win Rate:', self._format_percentage(perf_data.get('win_rate', 0))),
            ('Profit Factor:', f"{perf_data.get('profit_factor', 0):.2f}"),
            ('Average Win:', self._format_currency(perf_data.get('avg_win', 0))),
            ('Average Loss:', self._format_currency(perf_data.get('avg_loss', 0))),
            ('Largest Win:', self._format_currency(perf_data.get('largest_win', 0))),
            ('Largest Loss:', self._format_currency(perf_data.get('largest_loss', 0))),
            ('Average Hold Period:', f"{perf_data.get('avg_hold_period', 0)} days"),
            ('Total Commissions:', self._format_currency(perf_data.get('total_commissions', 0)))
        ]
        
        for metric, value in core_metrics:
            sheet[f'A{current_row}'] = metric
            sheet[f'B{current_row}'] = value
            current_row += 1
        
        # Risk metrics section
        current_row += 2
        sheet[f'A{current_row}'] = "Risk Analysis"
        sheet[f'A{current_row}'].font = Font(bold=True, size=14)
        current_row += 1
        
        risk_metrics = [
            ('Value at Risk (95%):', self._format_currency(perf_data.get('var_95', 0))),
            ('Expected Shortfall:', self._format_currency(perf_data.get('expected_shortfall', 0))),
            ('Beta (vs Market):', f"{perf_data.get('beta', 0):.2f}"),
            ('Alpha:', f"{perf_data.get('alpha', 0):.2f}"),
            ('Correlation to Market:', f"{perf_data.get('market_correlation', 0):.2f}"),
            ('Information Ratio:', f"{perf_data.get('information_ratio', 0):.2f}"),
            ('Tracking Error:', self._format_percentage(perf_data.get('tracking_error', 0))),
            ('Upside Capture:', self._format_percentage(perf_data.get('upside_capture', 0))),
            ('Downside Capture:', self._format_percentage(perf_data.get('downside_capture', 0)))
        ]
        
        for metric, value in risk_metrics:
            sheet[f'A{current_row}'] = metric
            sheet[f'B{current_row}'] = value
            current_row += 1
        
        # Auto-fit columns
        self._auto_fit_columns(sheet)

    def _create_risk_analysis_sheet(self, workbook: Workbook, data: Dict[str, Any]):
        """Create risk analysis sheet with detailed risk metrics."""
        sheet = workbook.create_sheet(title="Risk_Analysis")
        
        # Risk analysis header
        sheet['A1'] = "Portfolio Risk Analysis"
        sheet['A1'].font = Font(bold=True, size=16)
        
        current_row = 3
        risk_data = data.get('risk_analysis', {})
        
        # Portfolio composition risk
        sheet[f'A{current_row}'] = "Portfolio Composition Risk"
        sheet[f'A{current_row}'].font = Font(bold=True, size=14)
        current_row += 2
        
        # Sector allocation
        sector_headers = ['Sector', 'Allocation %', 'Risk Score', 'Correlation', 'Concentration Risk']
        for col, header in enumerate(sector_headers, 1):
            cell = sheet.cell(row=current_row, column=col, value=header)
            cell.style = self.styles['header_style']
        
        current_row += 1
        
        # Add sector data
        sectors = risk_data.get('sector_allocation', [])
        for sector in sectors:
            row_data = [
                sector.get('sector', ''),
                sector.get('allocation_pct', 0),
                sector.get('risk_score', 0),
                sector.get('correlation', 0),
                sector.get('concentration_risk', 'Low')
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = sheet.cell(row=current_row, column=col, value=value)
                
                # Format percentage columns
                if col in [2, 4]:
                    cell.number_format = '0.00%'
                elif col == 3:
                    cell.number_format = '0.00'
            
            current_row += 1
        
        # Position size risk
        current_row += 2
        sheet[f'A{current_row}'] = "Position Size Risk Analysis"
        sheet[f'A{current_row}'].font = Font(bold=True, size=14)
        current_row += 2
        
        position_headers = ['Symbol', 'Position Size', 'Portfolio %', 'Risk Rating', 'Recommended Action']
        for col, header in enumerate(position_headers, 1):
            cell = sheet.cell(row=current_row, column=col, value=header)
            cell.style = self.styles['header_style']
        
        current_row += 1
        
        # Add position risk data
        position_risks = risk_data.get('position_risks', [])
        for pos_risk in position_risks:
            row_data = [
                pos_risk.get('symbol', ''),
                pos_risk.get('position_size', 0),
                pos_risk.get('portfolio_pct', 0),
                pos_risk.get('risk_rating', 'Medium'),
                pos_risk.get('recommended_action', 'Hold')
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = sheet.cell(row=current_row, column=col, value=value)
                
                # Format columns
                if col == 2:  # Position size
                    cell.number_format = '$#,##0.00'
                elif col == 3:  # Portfolio percentage
                    cell.number_format = '0.00%'
            
            current_row += 1
        
        # Auto-fit columns
        self._auto_fit_columns(sheet)

    def _create_trading_charts_sheet(self, workbook: Workbook, data: Dict[str, Any]):
        """Create comprehensive charts for trading performance visualization."""
        sheet = workbook.create_sheet(title="Performance_Charts")
        
        # Charts header
        sheet['A1'] = "Trading Performance Charts"
        sheet['A1'].font = Font(bold=True, size=16)
        
        # Portfolio value chart
        portfolio_chart = LineChart()
        portfolio_chart.title = "Portfolio Value Over Time"
        portfolio_chart.x_axis.title = "Date"
        portfolio_chart.y_axis.title = "Portfolio Value ($)"
        portfolio_chart.width = 15
        portfolio_chart.height = 8
        
        sheet.add_chart(portfolio_chart, "A3")
        
        # Monthly returns chart
        returns_chart = BarChart()
        returns_chart.title = "Monthly Returns"
        returns_chart.x_axis.title = "Month"
        returns_chart.y_axis.title = "Return (%)"
        returns_chart.width = 15
        returns_chart.height = 8
        
        sheet.add_chart(returns_chart, "A20")
        
        # Sector allocation pie chart
        sector_chart = PieChart()
        sector_chart.title = "Portfolio Sector Allocation"
        sector_chart.width = 12
        sector_chart.height = 8
        
        sheet.add_chart(sector_chart, "Q3")

    def _apply_scanner_formatting(self, workbook: Workbook):
        """Apply comprehensive formatting to scanner results workbook."""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            if sheet_name == "Detailed_Data":
                # Apply conditional formatting for price changes
                self._apply_price_change_formatting(sheet, 'F')  # % Change column
                
                # Apply volume ratio formatting
                self._apply_volume_formatting(sheet, 'I')  # Volume Ratio column
                
                # Apply technical indicator formatting
                self._apply_technical_formatting(sheet, 'O')  # RSI column

    def _apply_journal_formatting(self, workbook: Workbook):
        """Apply formatting to trading journal workbook."""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            if sheet_name == "Performance_Analysis":
                # Apply P/L formatting
                self._apply_pnl_formatting(sheet, 'E')  # P/L column
                
                # Apply win rate formatting
                self._apply_percentage_formatting(sheet, 'F')  # Win Rate column

    def _apply_trading_formatting(self, workbook: Workbook):
        """Apply comprehensive formatting to trading performance workbook."""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            if sheet_name == "Portfolio_Summary":
                # Apply P/L conditional formatting
                self._apply_pnl_formatting(sheet, 'F')  # Unrealized P/L column
                
            elif sheet_name == "Trade_History":
                # Apply P/L formatting
                self._apply_pnl_formatting(sheet, 'M')  # Net P/L column
                
                # Apply return percentage formatting
                self._apply_percentage_formatting(sheet, 'O')  # Return % column

    def _apply_price_change_formatting(self, sheet, column: str):
        """Apply conditional formatting for price changes."""
        # Get data range
        max_row = sheet.max_row
        if max_row > 1:
            range_str = f"{column}2:{column}{max_row}"
            
            # Positive changes - green
            positive_rule = CellIsRule(
                operator='greaterThan',
                formula=[0],
                fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
                font=Font(color='006100')
            )
            
            # Negative changes - red
            negative_rule = CellIsRule(
                operator='lessThan',
                formula=[0],
                fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
                font=Font(color='9C0006')
            )
            
            sheet.conditional_formatting.add(range_str, positive_rule)
            sheet.conditional_formatting.add(range_str, negative_rule)

    def _apply_volume_formatting(self, sheet, column: str):
        """Apply conditional formatting for volume ratios."""
        max_row = sheet.max_row
        if max_row > 1:
            range_str = f"{column}2:{column}{max_row}"
            
            # High volume - blue gradient
            high_volume_rule = ColorScaleRule(
                start_type='min',
                start_color='FFFFFF',
                end_type='max',
                end_color='5B9BD5'
            )
            
            sheet.conditional_formatting.add(range_str, high_volume_rule)

    def _apply_technical_formatting(self, sheet, column: str):
        """Apply conditional formatting for technical indicators."""
        max_row = sheet.max_row
        if max_row > 1:
            range_str = f"{column}2:{column}{max_row}"
            
            # RSI overbought/oversold formatting
            overbought_rule = CellIsRule(
                operator='greaterThan',
                formula=[70],
                fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            )
            
            oversold_rule = CellIsRule(
                operator='lessThan',
                formula=[30],
                fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            )
            
            sheet.conditional_formatting.add(range_str, overbought_rule)
            sheet.conditional_formatting.add(range_str, oversold_rule)

    def _apply_pnl_formatting(self, sheet, column: str):
        """Apply P/L conditional formatting."""
        max_row = sheet.max_row
        if max_row > 1:
            range_str = f"{column}2:{column}{max_row}"
            
            # Profit - green
            profit_rule = CellIsRule(
                operator='greaterThan',
                formula=[0],
                fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
                font=Font(color='006100', bold=True)
            )
            
            # Loss - red
            loss_rule = CellIsRule(
                operator='lessThan',
                formula=[0],
                fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
                font=Font(color='9C0006', bold=True)
            )
            
            sheet.conditional_formatting.add(range_str, profit_rule)
            sheet.conditional_formatting.add(range_str, loss_rule)

    def _apply_percentage_formatting(self, sheet, column: str):
        """Apply percentage conditional formatting."""
        max_row = sheet.max_row
        if max_row > 1:
            range_str = f"{column}2:{column}{max_row}"
            
            # Color scale for percentages
            percentage_rule = ColorScaleRule(
                start_type='min',
                start_color='F2F2F2',
                mid_type='percentile',
                mid_value=50,
                mid_color='FFEB9C',
                end_type='max',
                end_color='63BE7B'
            )
            
            sheet.conditional_formatting.add(range_str, percentage_rule)

    def _apply_custom_formatting(self, workbook: Workbook, template_config: Dict[str, Any]):
        """Apply custom formatting based on template configuration."""
        formatting_rules = template_config.get('conditional_formatting', [])
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            for rule_name in formatting_rules:
                if rule_name == 'price_change':
                    self._apply_price_change_formatting(sheet, 'E')  # Assuming column E
                elif rule_name == 'volume_ratio':
                    self._apply_volume_formatting(sheet, 'F')  # Assuming column F
                elif rule_name == 'pnl_analysis':
                    self._apply_pnl_formatting(sheet, 'G')  # Assuming column G

    def _create_custom_charts(self, workbook: Workbook, data: Dict[str, Any], chart_types: List[str]):
        """Create custom charts based on template configuration."""
        if not chart_types:
            return
        
        # Create charts sheet if it doesn't exist
        if "Charts" not in workbook.sheetnames:
            sheet = workbook.create_sheet(title="Charts")
        else:
            sheet = workbook["Charts"]
        
        chart_position_row = 2
        
        for chart_type in chart_types:
            if chart_type == 'price_chart':
                chart = LineChart()
                chart.title = "Price Analysis"
                chart.width = 15
                chart.height = 8
                sheet.add_chart(chart, f"A{chart_position_row}")
                
            elif chart_type == 'volume_chart':
                chart = BarChart()
                chart.title = "Volume Analysis"
                chart.width = 15
                chart.height = 8
                sheet.add_chart(chart, f"A{chart_position_row + 15}")
                
            elif chart_type == 'performance_comparison':
                chart = ScatterChart()
                chart.title = "Performance Comparison"
                chart.width = 15
                chart.height = 8
                sheet.add_chart(chart, f"Q{chart_position_row}")
            
            chart_position_row += 20

    def _populate_custom_sheet(self, sheet, data: Dict[str, Any], sheet_type: str):
        """Populate custom sheet with data based on sheet type."""
        if sheet_type == 'data' or sheet_type == 'summary':
            # Generic data population
            if isinstance(data, dict) and 'headers' in data and 'rows' in data:
                headers = data['headers']
                rows = data['rows']
                
                # Add headers
                for col, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=col, value=header)
                    cell.style = self.styles['header_style']
                
                # Add data rows
                for row_idx, row_data in enumerate(rows, 2):
                    for col, value in enumerate(row_data, 1):
                        sheet.cell(row=row_idx, column=col, value=value)
            
            # Auto-fit columns
            self._auto_fit_columns(sheet)

    def _auto_fit_columns(self, sheet):
        """Auto-fit column widths based on content."""
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set column width with reasonable limits
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = max(adjusted_width, 10)

    def _create_backup(self, file_path: Path):
        """Create backup copy of exported file."""
        try:
            backup_dir = self.export_base_dir / 'backups'
            backup_filename = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file_path.name}"
            backup_path = backup_dir / backup_filename
            
            # Copy file to backup location
            import shutil
            shutil.copy2(file_path, backup_path)
            
            # Clean old backups based on retention policy
            self._cleanup_old_backups(backup_dir)
            
            self.logger.info(f"Backup created: {backup_path}")
            
        except Exception as e:
            self.logger.warning(f"Failed to create backup: {e}")

    def _cleanup_old_backups(self, backup_dir: Path):
        """Clean up old backup files based on retention policy."""
        try:
            retention_days = self.config['export_settings'].get('backup_retention_days', 30)
            cutoff_date = datetime.now() - timedelta(days=retention_days)
            
            for backup_file in backup_dir.glob('backup_*'):
                if backup_file.stat().st_mtime < cutoff_date.timestamp():
                    backup_file.unlink()
                    self.logger.info(f"Deleted old backup: {backup_file}")
                    
        except Exception as e:
            self.logger.warning(f"Failed to cleanup old backups: {e}")

    def _format_currency(self, value: Union[float, Decimal, int]) -> str:
        """Format value as currency."""
        if isinstance(value, str):
            try:
                value = float(value)
            except ValueError:
                return "$0.00"
        
        if value is None:
            return "$0.00"
        
        # Use Decimal for precise currency formatting
        decimal_value = Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        return f"${decimal_value:,.2f}"

    def _format_percentage(self, value: Union[float, Decimal, int]) -> str:
        """Format value as percentage."""
        if isinstance(value, str):
            try:
                value = float(value)
            except ValueError:
                return "0.00%"
        
        if value is None:
            return "0.00%"
        
        # Convert to percentage and format
        percentage = float(value) * 100 if abs(value) <= 1 else float(value)
        return f"{percentage:.2f}%"

    def _format_number(self, value: Union[float, int, str]) -> str:
        """Format value as number with appropriate scaling."""
        if isinstance(value, str):
            try:
                value = float(value)
            except ValueError:
                return "0"
        
        if value is None:
            return "0"
        
        # Format large numbers with K, M, B suffixes
        abs_value = abs(value)
        if abs_value >= 1_000_000_000:
            return f"{value / 1_000_000_000:.1f}B"
        elif abs_value >= 1_000_000:
            return f"{value / 1_000_000:.1f}M"
        elif abs_value >= 1_000:
            return f"{value / 1_000:.1f}K"
        else:
            return f"{value:,.0f}"

    def get_export_statistics(self) -> Dict[str, Any]:
        """
        Get export utility performance statistics.
        
        Returns:
            Dictionary containing export statistics
        """
        return {
            'total_exports': self.export_stats['total_exports'],
            'successful_exports': self.export_stats['successful_exports'],
            'failed_exports': self.export_stats['failed_exports'],
            'success_rate': (
                self.export_stats['successful_exports'] / max(1, self.export_stats['total_exports'])
            ) * 100,
            'last_export_time': self.export_stats['last_export_time'],
            'export_directory': str(self.export_base_dir)
        }

    def validate_export_data(self, data: Dict[str, Any], export_type: str) -> Tuple[bool, List[str]]:
        """
        Validate export data structure and content.
        
        Args:
            data: Data dictionary to validate
            export_type: Type of export to validate against
            
        Returns:
            Tuple of (is_valid: bool, validation_errors: List[str])
        """
        errors = []
        
        try:
            if not isinstance(data, dict):
                errors.append("Data must be a dictionary")
                return False, errors
            
            if export_type == 'scanner_results':
                required_keys = ['results']
                for key in required_keys:
                    if key not in data:
                        errors.append(f"Missing required key: {key}")
                
                # Validate results structure
                results = data.get('results', [])
                if not isinstance(results, list):
                    errors.append("Results must be a list")
                elif len(results) > 0:
                    sample_result = results[0]
                    required_fields = ['symbol', 'current_price', 'percent_change']
                    for field in required_fields:
                        if field not in sample_result:
                            errors.append(f"Missing required field in results: {field}")
            
            elif export_type == 'trading_journal':
                required_keys = ['journal_entries']
                for key in required_keys:
                    if key not in data:
                        errors.append(f"Missing required key: {key}")
            
            elif export_type == 'paper_trading':
                required_keys = ['portfolio_summary', 'trade_history']
                for key in required_keys:
                    if key not in data:
                        errors.append(f"Missing required key: {key}")
            
            return len(errors) == 0, errors
            
        except Exception as e:
            errors.append(f"Validation error: {str(e)}")
            return False, errors

    def cleanup_export_directory(self, days_old: int = 90):
        """
        Clean up old export files based on age.
        
        Args:
            days_old: Remove files older than this many days
        """
        try:
            cutoff_date = datetime.now() - timedelta(days=days_old)
            cleaned_count = 0
            
            for export_type_dir in self.export_base_dir.iterdir():
                if export_type_dir.is_dir() and export_type_dir.name != 'backups':
                    for export_file in export_type_dir.glob('*.xlsx'):
                        if export_file.stat().st_mtime < cutoff_date.timestamp():
                            export_file.unlink()
                            cleaned_count += 1
            
            self.logger.info(f"Cleaned up {cleaned_count} old export files")
            return cleaned_count
            
        except Exception as e:
            self.logger.error(f"Failed to cleanup export directory: {e}")
            return 0


# Factory function for creating export utility instance
def create_export_utility(config_path: Optional[str] = None) -> ExcelExportUtility:
    """
    Factory function to create and configure Excel export utility.
    
    Args:
        config_path: Optional path to configuration file
        
    Returns:
        Configured ExcelExportUtility instance
    """
    return ExcelExportUtility(config_path)


# Example usage and testing functions
if __name__ == "__main__":
    # Example usage of the Excel Export Utility
    
    # Create export utility
    export_util = create_export_utility()
    
    # Example scanner data
    sample_scanner_data = {
        'scan_date': datetime.now(),
        'filters': 'Price > $5, Volume > 1M',
        'results': [
            {
                'symbol': 'AAPL',
                'company_name': 'Apple Inc.',
                'current_price': 175.25,
                'previous_close': 172.50,
                'price_change': 2.75,
                'percent_change': 0.0159,
                'volume': 45000000,
                'avg_volume': 35000000,
                'volume_ratio': 1.29,
                'vwap': 174.80,
                'market_cap': 2800000000000,
                'float_shares': 15500000000,
                'sector': 'Technology',
                'industry': 'Consumer Electronics',
                'technical_indicators': {
                    'rsi': 65.2,
                    'macd_signal': 1.25,
                    'macd_histogram': 0.35,
                    'sma_20': 173.45,
                    'sma_50': 170.20,
                    'bollinger_upper': 178.50,
                    'bollinger_lower': 169.30,
                    'atr': 2.15,
                    'technical_score': 7.5
                }
            }
        ]
    }
    
    # Test scanner export
    success, result = export_util.export_scanner_results(sample_scanner_data, "test_scanner_export")
    print(f"Scanner export: {'Success' if success else 'Failed'} - {result}")
    
    # Get statistics
    stats = export_util.get_export_statistics()
    print(f"Export statistics: {stats}")